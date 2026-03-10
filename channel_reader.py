"""
Модуль для чтения новостей из публичных Telegram каналов через Client API
"""
import logging
import asyncio
from datetime import datetime, timedelta, timezone
from typing import List, Dict, Optional
from telethon import TelegramClient
from telethon.tl.types import Channel, Message
import sqlite3
import os
from config import MONITORED_CHANNELS, DATABASE_URL

logger = logging.getLogger(__name__)


class ChannelReader:
    """Класс для чтения новостей из публичных Telegram каналов"""
    
    def __init__(self, api_id: str, api_hash: str, phone: Optional[str] = None):
        """
        Инициализация клиента Telegram
        
        Args:
            api_id: API ID от https://my.telegram.org
            api_hash: API Hash от https://my.telegram.org
            phone: Номер телефона (опционально, для авторизации)
        """
        self.api_id = api_id
        self.api_hash = api_hash
        self.phone = phone
        self.client = None
        self.monitored_channels = MONITORED_CHANNELS
        
        # Обрабатываем путь к базе данных
        if DATABASE_URL.startswith("sqlite:///"):
            self.db_path = DATABASE_URL.replace("sqlite:///", "")
        else:
            self.db_path = DATABASE_URL
        if not os.path.isabs(self.db_path):
            self.db_path = os.path.join(os.path.dirname(__file__), self.db_path)
    
    async def connect(self):
        """Подключение к Telegram"""
        if not self.client:
            session_name = "telegram_session"
            self.client = TelegramClient(session_name, self.api_id, self.api_hash)
        
        # Подключаемся к серверу
        try:
            if not self.client.is_connected():
                await self.client.connect()
        except:
            # Если еще не подключены, подключаемся
            await self.client.connect()
        
        # Проверяем авторизацию
        if not await self.client.is_user_authorized():
            # Требуется авторизация
            if hasattr(self, '_auth_in_progress') and self._auth_in_progress:
                # Уже идет процесс авторизации
                return
            
            self._auth_in_progress = True
            
            # Проверяем, интерактивный ли это запуск (есть ли stdin)
            import sys
            if sys.stdin.isatty():
                # Интерактивный режим - можно запрашивать код и пароль
                print("🔐 Авторизация в Telegram...")
                print("📱 Вам придет код подтверждения в Telegram")
                
                await self.client.send_code_request(self.phone)
                code = input('📨 Введите код подтверждения из Telegram: ')
                
                try:
                    await self.client.sign_in(self.phone, code)
                    print("✅ Авторизация успешна!")
                except Exception as e:
                    # Если требуется пароль 2FA
                    error_str = str(e).lower()
                    if "password" in error_str or "PASSWORD" in str(e):
                        print("🔒 Требуется пароль двухфакторной аутентификации (2FA)")
                        password = input('🔒 Введите пароль 2FA: ')
                        await self.client.sign_in(password=password)
                        print("✅ Авторизация успешна!")
                    else:
                        raise
            else:
                # Неинтерактивный режим (из бота) - используем сохраненную сессию
                logger.warning("Требуется авторизация, но режим неинтерактивный. Запустите channel_reader.py отдельно.")
                raise Exception("Требуется авторизация. Запустите 'python channel_reader.py' для первой авторизации.")
        
        logger.info("Подключено к Telegram через Client API")
    
    async def disconnect(self):
        """Отключение от Telegram"""
        if self.client and self.client.is_connected():
            await self.client.disconnect()
            logger.info("Отключено от Telegram")
    
    async def fetch_channel_messages(self, channel_username: str, limit: int = 100, 
                                     hours: int = 24) -> List[Dict]:
        """
        Получение сообщений из канала
        
        Args:
            channel_username: Username канала (например: @rosavtodor)
            limit: Максимальное количество сообщений
            hours: За какой период получать сообщения
        
        Returns:
            Список сообщений
        """
        if not self.client:
            await self.connect()
        
        # Убеждаемся, что подключены и авторизованы
        if not self.client.is_connected():
            await self.client.connect()
        
        if not await self.client.is_user_authorized():
            raise Exception("Не авторизован. Запустите 'python channel_reader.py' для авторизации.")
        
        try:
            # Получаем канал
            entity = await self.client.get_entity(channel_username)
            
            if not isinstance(entity, Channel):
                logger.warning(f"{channel_username} не является каналом")
                return []
            
            # Граница периода: только сообщения не старше hours
            now_utc = datetime.now(timezone.utc)
            since_date = now_utc - timedelta(hours=hours)
            
            # Берём с запасом, чтобы после фильтра по дате осталось до limit сообщений за hours
            scan_limit = max(limit, 200)
            messages = []
            count = 0
            async for message in self.client.iter_messages(entity, limit=scan_limit):
                if not (isinstance(message, Message) and (message.text or message.message)):
                    continue
                msg_date = message.date
                if msg_date.tzinfo is None:
                    msg_date = msg_date.replace(tzinfo=timezone.utc)
                if msg_date < since_date:
                    continue
                messages.append({
                    "id": message.id,
                    "date": message.date,
                    "text": message.text or message.message or "",
                    "chat_id": entity.id,
                    "chat_username": channel_username
                })
                count += 1
                if count % 5 == 0:
                    print(f"   📥 Обработано сообщений: {count}...", end='\r')
            
            print(f"   📥 Обработано сообщений: {count}        ")
            logger.info(f"Получено {len(messages)} сообщений из {channel_username} (за последние {hours} ч)")
            return messages
            
        except Exception as e:
            logger.error(f"Ошибка при получении сообщений из {channel_username}: {e}")
            return []
    
    def save_news_to_db(self, messages: List[Dict], source_name: str):
        """
        Сохранение новостей в базу данных
        
        Args:
            messages: Список сообщений
            source_name: Название источника
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        saved_count = 0
        for msg in messages:
            try:
                text = msg.get("text", "")
                if not text:
                    continue
                
                # Извлекаем заголовок
                title = None
                if text:
                    lines = text.split("\n")
                    if len(lines) > 1:
                        title = lines[0][:200]
                
                # Дата появления новости в канале (дата публикации сообщения)
                msg_date = msg.get("date")
                if isinstance(msg_date, datetime):
                    if msg_date.tzinfo is not None:
                        # приводим к локальному времени и сохраняем как naive datetime
                        date_val = msg_date.astimezone().replace(tzinfo=None)
                    else:
                        date_val = msg_date
                else:
                    date_val = datetime.now()
                
                cursor.execute("""
                    INSERT OR IGNORE INTO news 
                    (message_id, chat_id, chat_username, source_name, title, text, date)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    msg["id"],
                    msg["chat_id"],
                    msg["chat_username"],
                    source_name,
                    title,
                    text,
                    date_val
                ))
                saved_count += 1
            except Exception as e:
                logger.error(f"Ошибка при сохранении новости: {e}")
        
        conn.commit()
        conn.close()
        if saved_count > 0:
            print(f"   💾 Сохранено в БД: {saved_count} новостей")
        logger.info(f"Сохранено {saved_count} новостей из {source_name}")

    def save_news_to_excel(self, news_list: List[Dict], filepath: Optional[str] = None) -> str:
        """
        Записывает список новостей в Excel-файл.
        Каждая запись: source, title, text, date.

        Args:
            news_list: Список dict с ключами source, title, text, date
            filepath: Путь к файлу (если None — создаётся в папке скрипта с именем novosti_YYYY-MM-DD_HH-MM.xlsx)

        Returns:
            Путь к созданному файлу
        """
        if not news_list:
            logger.warning("save_news_to_excel: пустой список")
            return ""
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("Для выгрузки в Excel установите openpyxl: pip install openpyxl")
            return ""
        if filepath is None:
            filepath = os.path.join(
                os.path.dirname(__file__),
                "novosti_" + datetime.now().strftime("%Y-%m-%d_%H-%M") + ".xlsx"
            )
        wb = Workbook()
        ws = wb.active
        ws.title = "Новости"
        headers = ["Отрасль", "Источник", "Заголовок", "Текст", "Дата"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, item in enumerate(news_list, 2):
            ws.cell(row=row_idx, column=1, value=item.get("industry", ""))
            ws.cell(row=row_idx, column=2, value=item.get("source", ""))
            ws.cell(row=row_idx, column=3, value=item.get("title", ""))
            ws.cell(row=row_idx, column=4, value=item.get("text", ""))
            date_val = item.get("date", "")
            ws.cell(row=row_idx, column=5, value=str(date_val) if date_val else "")
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        ws.column_dimensions["D"].width = 50
        wb.save(filepath)
        logger.info(f"Выгрузка в Excel: {len(news_list)} записей → {filepath}")
        return filepath

    async def fetch_all_channels(self, hours: int = 24, limit_per_channel: int = 50):
        """
        Получение новостей из всех отслеживаемых каналов
        
        Args:
            hours: За какой период получать новости
            limit_per_channel: Максимальное количество сообщений на канал
        """
        await self.connect()
        
        total_fetched = 0
        for source_name, channel_username in self.monitored_channels.items():
            if not channel_username or not channel_username.startswith("@"):
                continue
            
            logger.info(f"Получение новостей из {source_name} ({channel_username})...")
            messages = await self.fetch_channel_messages(
                channel_username,
                limit=limit_per_channel,
                hours=hours
            )
            
            if messages:
                self.save_news_to_db(messages, source_name)
                total_fetched += len(messages)
            
            # Небольшая задержка между каналами
            await asyncio.sleep(2)
        
        logger.info(f"Всего получено новостей: {total_fetched}")
        return total_fetched


async def main():
    """Тестовая функция для проверки работы"""
    import os
    from dotenv import load_dotenv
    
    load_dotenv()
    
    api_id = os.getenv("TELEGRAM_API_ID")
    api_hash = os.getenv("TELEGRAM_API_HASH")
    phone = os.getenv("TELEGRAM_PHONE")
    
    print("\n" + "="*70)
    print("🚀 ЗАПУСК СБОРЩИКА НОВОСТЕЙ ИЗ TELEGRAM КАНАЛОВ")
    print("="*70)
    
    if not api_id or not api_hash:
        print("\n❌ ОШИБКА: Не указаны TELEGRAM_API_ID и TELEGRAM_API_HASH в .env")
        print("Получите их на https://my.telegram.org")
        return
    
    if not phone:
        print("\n❌ ОШИБКА: Не указан TELEGRAM_PHONE в .env")
        print("Укажите ваш номер телефона в формате +7XXXXXXXXXX")
        return
    
    print(f"\n📋 Настройки:")
    print(f"   API ID: {api_id}")
    print(f"   API Hash: {api_hash[:10]}...")
    print(f"   Телефон: {phone}")
    
    reader = ChannelReader(api_id, api_hash, phone)
    
    try:
        print("\n🔌 Подключение к Telegram...")
        await reader.connect()
        print("✅ Подключено успешно!\n")

        # Спрашиваем период выгрузки
        print("За какой период выгрузить новости?")
        print("  1 — последние 24 часа (1 день)")
        print("  7 — последняя неделя (7 дней)")
        print("  30 — последний месяц (30 дней)")
        print("  Или введите число дней (например 3):")
        period_input = input("Период (по умолчанию 1): ").strip() or "1"
        if period_input == "1":
            hours = 24
            period_label = "последние 24 часа"
        elif period_input == "7":
            hours = 168
            period_label = "последняя неделя"
        elif period_input == "30":
            hours = 720
            period_label = "последний месяц"
        else:
            try:
                days = int(period_input)
                if days < 1:
                    days = 1
                elif days > 365:
                    days = 365
                hours = days * 24
                period_label = f"последние {days} дн."
            except ValueError:
                hours = 24
                period_label = "последние 24 часа"
                print(f"   ⚠️ Неверный ввод, используется период: {period_label}")
        limit_per_channel = 50
        print(f"\n📅 Период выгрузки: {period_label} (до {limit_per_channel} сообщений на канал)\n")
        
        # Получаем новости по всем отраслям и общим каналам
        from config import INDUSTRY_SOURCES, INDUSTRY_DISPLAY_NAMES, GENERAL_SOURCES
        
        total_channels = sum(len(s) for s in INDUSTRY_SOURCES.values() if s) + len(GENERAL_SOURCES)
        if total_channels == 0:
            print("⚠️  Источники по отраслям не настроены в config.py (INDUSTRY_SOURCES)")
            return
        
        print(f"📰 Найдено источников по отраслям: {total_channels}")
        print("-" * 70)

        all_news = []
        total_saved = 0
        for industry, sources in INDUSTRY_SOURCES.items():
            if not sources:
                continue
            display = INDUSTRY_DISPLAY_NAMES.get(industry, industry)
            for source_name, channel_username in sources.items():
                if channel_username:
                    print(f"\n📡 [{display}] {source_name} ({channel_username})")
                    print(f"   Период: {period_label}, лимит: {limit_per_channel} сообщений")
                    
                    messages = await reader.fetch_channel_messages(channel_username, limit=limit_per_channel, hours=hours)
                    
                    if messages:
                        print(f"   ✅ Получено: {len(messages)}, сохранено в БД")
                        reader.save_news_to_db(messages, source_name)
                        total_saved += len(messages)
                        for msg in messages:
                            text = msg.get("text", "") or ""
                            lines = text.split("\n")
                            title = (lines[0][:200] if lines else "") or "Без заголовка"
                            all_news.append({
                                "industry": display,
                                "source": source_name,
                                "title": title,
                                "text": text[:5000],
                                "date": msg.get("date", ""),
                            })
                    else:
                        print(f"   ⚠️  Новых сообщений не найдено")
                    
                    print("-" * 70)
                    await asyncio.sleep(1)

        # Общие каналы
        if GENERAL_SOURCES:
            print("\n📌 Общие каналы")
            print("-" * 70)
            for source_name, channel_username in GENERAL_SOURCES.items():
                if channel_username:
                    print(f"\n📡 [Общие] {source_name} ({channel_username})")
                    print(f"   Период: {period_label}, лимит: {limit_per_channel} сообщений")
                    messages = await reader.fetch_channel_messages(channel_username, limit=limit_per_channel, hours=hours)
                    if messages:
                        print(f"   ✅ Получено: {len(messages)}, сохранено в БД")
                        reader.save_news_to_db(messages, source_name)
                        total_saved += len(messages)
                        for msg in messages:
                            text = msg.get("text", "") or ""
                            lines = text.split("\n")
                            title = (lines[0][:200] if lines else "") or "Без заголовка"
                            all_news.append({
                                "industry": "Общие",
                                "source": source_name,
                                "title": title,
                                "text": text[:5000],
                                "date": msg.get("date", ""),
                            })
                    else:
                        print(f"   ⚠️  Новых сообщений не найдено")
                    print("-" * 70)
                    await asyncio.sleep(1)

        if all_news:
            excel_path = reader.save_news_to_excel(all_news)
            if excel_path:
                print(f"\n📊 Все найденные новости записаны в Excel: {excel_path}")
        
        print("\n" + "="*70)
        print(f"✅ СБОР НОВОСТЕЙ ЗАВЕРШЕН")
        print(f"📊 Всего сохранено новостей: {total_saved}")
        print("="*70)
        
        if total_saved > 0:
            print("\n💡 Запустите бота (python bot.py) и используйте:")
            print("   /summary_topic <отрасль> — сводка по отрасли")
            print("   /topics — список отраслей")
            print("   /stats — статистика")
        else:
            print("\n⚠️  Новости не найдены. Возможные причины:")
            print(f"   - В каналах нет новых сообщений за выбранный период ({period_label})")
            print("   - Каналы недоступны или приватные")
            print("   - Проверьте правильность username каналов в config.py")
        
        print("\n")
        
    except Exception as e:
        print(f"\n❌ ОШИБКА: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        print("🔌 Отключение от Telegram...")
        await reader.disconnect()
        print("✅ Отключено\n")


if __name__ == "__main__":
    asyncio.run(main())