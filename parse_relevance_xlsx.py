# Скрипт для анализа Excel: извлечение строк с зелёной заливкой (релевантные новости по отраслям)
import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

def _safe(s, maxlen=500):
    return re.sub(r'[^\w\s.,;:!?\-—–()\"\'%№@#$^&*+=/\\\n\r\t\u0400-\u04FF]', '', s)[:maxlen]

def is_green_fill(fill):
    if not fill or getattr(fill, "fill_type", None) != "solid":
        return False
    fg = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
    if not fg:
        return False
    # Theme color: в Excel зелёная заливка часто theme index 4 (зелёный акцент)
    if getattr(fg, "type", None) == "theme":
        theme = getattr(fg, "theme", 0)
        return theme in (3, 4, 5)  # типичные зелёные в теме
    rgb = getattr(fg, "rgb", None)
    if not rgb or len(str(rgb)) < 6:
        return False
    rgb = str(rgb).upper().replace("FF", "")[-6:]
    if len(rgb) != 6:
        return False
    try:
        r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return False
    return g > 100 and g > r and g > b

def main():
    path = Path(r"c:\Users\Admin\Downloads\novosti_2026-02-24_12-51.xlsx")
    if not path.exists():
        print("File not found: " + str(path))
        sys.exit(1)
    wb = openpyxl.load_workbook(path, data_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
            row_has_green = any(
                cell.fill and is_green_fill(cell.fill)
                for cell in row
            )
            if row_has_green:
                row_text = " | ".join(str(c.value or "").strip() for c in row if c.value)
                if row_text.strip():
                    results.append((sheet_name, row_idx, row_text[:600]))
    wb.close()
    out_path = path.parent / "relevance_analysis.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        by_sheet = {}
        for sheet, row_num, text in results:
            by_sheet.setdefault(sheet, []).append((row_num, text))
        for sheet in sorted(by_sheet.keys()):
            f.write(f"\n=== Отрасль (лист): {sheet} ===\n")
            for row_num, text in sorted(by_sheet[sheet], key=lambda x: x[0]):
                snip = text[:400] + ("..." if len(text) > 400 else "")
                f.write(f"  Строка {row_num}: {snip}\n")
        f.write("\n--- Примеры релевантных формулировок (все зелёные строки) ---\n")
        for i, (sheet, _, t) in enumerate(results[:80], 1):
            f.write(f"{i}. [{sheet}] {_safe(t, 350)}\n")
    print("Written:", out_path)
    return results

if __name__ == "__main__":
    main()
