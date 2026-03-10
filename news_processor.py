"""
Модуль для обработки новостей и генерации сводок с помощью GigaChat
"""
import logging
import os
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from news_collector import NewsCollector
from gigachat_client import GigaChatClient
from config import NEWS_TOPICS, INDUSTRY_SOURCES, GENERAL_SOURCES

logger = logging.getLogger(__name__)


class NewsProcessor:
    """Класс для обработки новостей и генерации сводок"""
    
    def __init__(self):
        self.collector = NewsCollector()
        self.gigachat = GigaChatClient()
        self._last_sources = []  # Храним последний список источников

    def _industry_labels_for_source(self, source_name: str) -> str:
        labels: List[str] = []
        for industry, sources in INDUSTRY_SOURCES.items():
            if sources and source_name in sources:
                labels.append(industry)
        if GENERAL_SOURCES and source_name in GENERAL_SOURCES:
            labels.append("общие")
        return ", ".join(labels) if labels else ""

    def _decorate_items_with_industry(self, items: List[Dict]) -> List[Dict]:
        for item in items:
            src = item.get("source_name", "") or ""
            item["industry"] = self._industry_labels_for_source(src)
        return items

    def _telegram_message_link(self, item: Dict) -> str:
        """
        Формирует прямую ссылку на сообщение в Telegram.
        - Для публичных каналов: https://t.me/<username>/<message_id>
        - Для каналов без username: https://t.me/c/<internal_id>/<message_id>, где internal_id = chat_id без -100
        """
        msg_id = item.get("message_id")
        if not msg_id:
            return ""
        chat_username = (item.get("chat_username") or "").strip()
        if chat_username:
            username = chat_username[1:] if chat_username.startswith("@") else chat_username
            return f"https://t.me/{username}/{msg_id}"
        chat_id = item.get("chat_id")
        try:
            chat_id_int = int(chat_id)
        except Exception:
            return ""
        # Telegram internal id for t.me/c links is chat_id without -100 prefix
        internal = str(abs(chat_id_int))
        if internal.startswith("100"):
            internal = internal[3:]
        return f"https://t.me/c/{internal}/{msg_id}"
    
    def generate_daily_summary(self, topic: Optional[str] = None, limit_sources: Optional[int] = None) -> str:
        """
        Генерация дневной сводки новостей
        
        Args:
            topic: Тема для фильтрации (опционально)
            limit_sources: Ограничение количества источников (опционально)
        
        Returns:
            Текст сводки
        """
        try:
            # Получаем новости за последние 24 часа
            if topic:
                # По каждой отрасли — свои источники; если есть — берём новости из них, иначе по теме
                sources = INDUSTRY_SOURCES.get(topic, {})
                if sources:
                    source_names = list(sources.keys())
                    news_items = self.collector.get_news_by_source(
                        source_names,
                        hours=168
                    )
                    logger.info(f"Найдено новостей по отрасли '{topic}' из {len(source_names)} источников: {len(news_items)}")
                else:
                    news_items = self.collector.get_news_by_topic(topic, hours=24)
            else:
                news_items = self.collector.get_recent_news(hours=24, limit=50)
            
            if not news_items:
                topic_text = f" по теме '{topic}'" if topic else ""
                return f"За последние 24 часа новостей{topic_text} не найдено. Попробуйте позже или выберите другую тему."
            
            # Если указано ограничение источников (когда у отрасли есть свои каналы)
            if limit_sources and INDUSTRY_SOURCES.get(topic):
                # Группируем новости по источникам
                sources_dict = {}
                for item in news_items:
                    source = item["source_name"]
                    if source not in sources_dict:
                        sources_dict[source] = []
                    sources_dict[source].append(item)
                
                # Берем только первые N источников, но из каждого источника берем несколько новостей
                # Это гарантирует, что в сводке будут представлены разные источники
                limited_items = []
                sources_list = list(sources_dict.items())[:limit_sources]
                
                # Из каждого источника берем примерно равное количество новостей
                max_news_per_source = max(4, 20 // limit_sources)  # Минимум 4 новости на источник
                
                for source, items in sources_list:
                    # Берем последние новости из каждого источника (они отсортированы по дате DESC)
                    limited_items.extend(items[:max_news_per_source])
                
                # Сортируем по дате (новые первыми) и ограничиваем общее количество
                limited_items.sort(key=lambda x: x.get("date", ""), reverse=True)
                news_items = limited_items[:20]  # Ограничиваем общее количество новостей
                
                used_sources = [source for source, _ in sources_list]
                logger.info(f"Использовано источников: {len(used_sources)} (ограничено до {limit_sources}): {used_sources}")
                logger.info(f"Новостей по источникам: {[(s, len([i for i in news_items if i['source_name'] == s])) for s in used_sources]}")
            
            # Формируем данные для GigaChat
            news_data = [
                {
                    "source": item["source_name"],
                    "title": item["title"],
                    "text": item["text"][:1000]  # Ограничиваем длину текста
                }
                for item in news_items
            ]
            
            # Получаем список уникальных источников
            sources_list = list(set([item["source_name"] for item in news_items]))
            
            # Генерируем сводку
            summary = self.gigachat.generate_summary(news_data, topic)
            
            # Отмечаем новости как обработанные
            news_ids = [item["id"] for item in news_items]
            self.collector.mark_as_processed(news_ids)
            
            # Сохраняем список источников для форматирования
            self._last_sources = sources_list
            
            return summary
            
        except Exception as e:
            logger.error(f"Ошибка при генерации сводки: {e}")
            return f"Произошла ошибка при генерации сводки: {str(e)}"
    
    def get_news_and_thesis_summary(
        self, topic: Optional[str] = None, limit_sources: Optional[int] = None
    ) -> Tuple[str, List[Dict[str, str]]]:
        """
        Получить новости и сгенерировать краткую тезисную сводку (нумерованный список).
        Новости не помечаются как обработанные. Список news_data возвращается для выбора пользователем.
        
        Returns:
            (thesis_text, news_data_list) — текст тезисной сводки и список dict с ключами source, title, text.
            При отсутствии новостей — (message_error, []).
        """
        try:
            if topic:
                sources = INDUSTRY_SOURCES.get(topic, {})
                if sources:
                    news_items = self.collector.get_news_by_source(
                        list(sources.keys()), hours=168
                    )
                    logger.info(f"Найдено новостей по отрасли '{topic}': {len(news_items)}")
                else:
                    news_items = self.collector.get_news_by_topic(topic, hours=24)
            else:
                news_items = self.collector.get_recent_news(hours=24, limit=50)
            
            if not news_items:
                topic_text = f" по теме '{topic}'" if topic else ""
                return (
                    f"За последние 24 часа новостей{topic_text} не найдено. Попробуйте позже или выберите другую тему.",
                    [],
                )
            
            if limit_sources and INDUSTRY_SOURCES.get(topic):
                sources_dict = {}
                for item in news_items:
                    s = item["source_name"]
                    if s not in sources_dict:
                        sources_dict[s] = []
                    sources_dict[s].append(item)
                limited_items = []
                for source, items in list(sources_dict.items())[:limit_sources]:
                    max_per = max(4, 20 // limit_sources)
                    limited_items.extend(items[:max_per])
                limited_items.sort(key=lambda x: x.get("date", ""), reverse=True)
                news_items = limited_items[:20]
            
            news_items = self.gigachat.filter_relevant_news(news_items)
            if not news_items:
                topic_text = f" по теме '{topic}'" if topic else ""
                return (
                    f"Релевантных новостей{topic_text} не найдено. Попробуйте другую тему или позже.",
                    [],
                )
            
            news_data = [
                {
                    "source": item["source_name"],
                    "title": item["title"],
                    "text": item["text"][:1000],
                }
                for item in news_items[:15]
            ]
            
            thesis_text = self.gigachat.generate_thesis_summary(news_data, topic)
            self._last_sources = list(set(item["source_name"] for item in news_items))
            return (thesis_text, news_data)
        except Exception as e:
            logger.error(f"Ошибка при получении тезисной сводки: {e}")
            return (f"Произошла ошибка: {str(e)}", [])

    def get_news_and_thesis_summary_by_sources(
        self, source_names: List[str], hours: int = 24
    ) -> Tuple[str, List[Dict[str, str]]]:
        """
        Общая тезисная сводка по конкретному списку источников.

        Используется, когда отрасль не выбрана (общие каналы).
        """
        try:
            news_items = self.collector.get_news_by_source(source_names, hours=hours)

            if not news_items:
                return ("За последние 24 часа новостей по общим каналам не найдено. Попробуйте позже.", [])

            news_items = self.gigachat.filter_relevant_news(news_items)
            if not news_items:
                return ("Релевантных новостей по общим каналам не найдено. Попробуйте позже.", [])

            news_data = [
                {
                    "source": item["source_name"],
                    "title": item["title"],
                    "text": item["text"][:1000],
                }
                for item in news_items[:15]
            ]

            thesis_text = self.gigachat.generate_thesis_summary(news_data, topic=None)
            self._last_sources = list(set(item["source_name"] for item in news_items))
            return (thesis_text, news_data)
        except Exception as e:
            logger.error(f"Ошибка при получении общей тезисной сводки: {e}")
            return (f"Произошла ошибка: {str(e)}", [])

    def export_relevant_news_to_excel(
        self, hours: int = 24, filepath: Optional[str] = None
    ) -> str:
        """
        Выгружает релевантные новости в Excel.
        Берёт новости из БД (после channel_reader / обработки каналов), за указанный период,
        фильтрует релевантные через GigaChat и сохраняет в Excel. В колонке «Дата» — дата появления в канале.

        Args:
            hours: За какой период брать новости (по умолчанию 24 часа).
            filepath: Путь к файлу (если None — создаётся relevant_novosti_YYYY-MM-DD_HH-MM.xlsx).

        Returns:
            Путь к созданному файлу или пустая строка при ошибке.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("Для выгрузки в Excel установите openpyxl: pip install openpyxl")
            return ""

        news_items = self.collector.get_recent_news(hours=hours, limit=5000)
        if not news_items:
            logger.warning("export_relevant_news_to_excel: новостей за период не найдено")
            return ""

        self._decorate_items_with_industry(news_items)
        relevant = self.gigachat.filter_relevant_news(news_items)

        if not relevant:
            logger.warning("export_relevant_news_to_excel: релевантных новостей не найдено")
            return ""

        if filepath is None:
            filepath = os.path.join(
                os.path.dirname(__file__),
                "relevant_novosti_" + datetime.now().strftime("%Y-%m-%d_%H-%M") + ".xlsx",
            )
        wb = Workbook()
        ws = wb.active
        ws.title = "Релевантные"
        headers = ["Отрасль", "Источник", "Ссылка", "Заголовок", "Текст", "Дата"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, item in enumerate(relevant, 2):
            ws.cell(row=row_idx, column=1, value=item.get("industry", ""))
            ws.cell(row=row_idx, column=2, value=item.get("source_name", ""))
            ws.cell(row=row_idx, column=3, value=self._telegram_message_link(item))
            ws.cell(row=row_idx, column=4, value=item.get("title", ""))
            ws.cell(row=row_idx, column=5, value=item.get("text", ""))
            ws.cell(row=row_idx, column=6, value=str(item.get("date", "")))
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["E"].width = 50
        wb.save(filepath)
        logger.info(f"Выгрузка релевантных в Excel: {len(relevant)} записей → {filepath}")
        return filepath

    def export_all_news_to_excel(self, hours: int = 24, filepath: Optional[str] = None) -> str:
        """
        Выгружает ВСЕ новости за период в Excel (без фильтра релевантности).
        В колонке «Дата» — дата появления в канале.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("Для выгрузки в Excel установите openpyxl: pip install openpyxl")
            return ""

        news_items = self.collector.get_recent_news(hours=hours, limit=5000)
        if not news_items:
            logger.warning("export_all_news_to_excel: новостей за период не найдено")
            return ""

        self._decorate_items_with_industry(news_items)

        if filepath is None:
            filepath = os.path.join(
                os.path.dirname(__file__),
                "all_novosti_" + datetime.now().strftime("%Y-%m-%d_%H-%M") + ".xlsx",
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Все"
        headers = ["Отрасль", "Источник", "Ссылка", "Заголовок", "Текст", "Дата"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, item in enumerate(news_items, 2):
            ws.cell(row=row_idx, column=1, value=item.get("industry", ""))
            ws.cell(row=row_idx, column=2, value=item.get("source_name", ""))
            ws.cell(row=row_idx, column=3, value=self._telegram_message_link(item))
            ws.cell(row=row_idx, column=4, value=item.get("title", ""))
            ws.cell(row=row_idx, column=5, value=item.get("text", ""))
            ws.cell(row=row_idx, column=6, value=str(item.get("date", "")))
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["E"].width = 50
        wb.save(filepath)
        logger.info(f"Выгрузка всех новостей в Excel: {len(news_items)} записей → {filepath}")
        return filepath

    def generate_infopovod_for_news(
        self, news_item: Dict[str, str], topic: Optional[str] = None
    ) -> str:
        """
        Сгенерировать полный инфоповод по одной новости (структура из промпта).
        
        Args:
            news_item: dict с ключами source, title, text
            topic: тема (опционально)
        
        Returns:
            Текст полного инфоповода.
        """
        return self.gigachat.generate_infopovod(news_item, topic=topic)
    
    def generate_topic_summaries(self) -> Dict[str, str]:
        """
        Генерация сводок по всем темам
        
        Returns:
            Словарь {тема: сводка}
        """
        summaries = {}
        
        for topic in NEWS_TOPICS:
            try:
                logger.info(f"Генерация сводки по теме: {topic}")
                summary = self.generate_daily_summary(topic=topic)
                summaries[topic] = summary
                # Очищаем источники после каждой генерации
                self._last_sources = []
            except Exception as e:
                logger.error(f"Ошибка при генерации сводки по теме {topic}: {e}")
                summaries[topic] = f"Не удалось сгенерировать сводку по теме '{topic}'"
        
        return summaries
    
    def categorize_and_save_topics(self, news_id: int, news_text: str):
        """
        Категоризация новости и сохранение тем в базу данных
        
        Args:
            news_id: ID новости в базе данных
            news_text: Текст новости
        """
        try:
            # Получаем темы через GigaChat
            topics = self.gigachat.categorize_news(news_text)
            
            if topics:
                # Сохраняем темы в базу данных
                import sqlite3
                
                conn = sqlite3.connect(self.collector.db_path)
                cursor = conn.cursor()
                
                topics_str = ", ".join(topics)
                cursor.execute("""
                    UPDATE news
                    SET topics = ?
                    WHERE id = ?
                """, (topics_str, news_id))
                
                conn.commit()
                conn.close()
                
                logger.info(f"Темы сохранены для новости {news_id}: {topics_str}")
        
        except Exception as e:
            logger.error(f"Ошибка при категоризации новости {news_id}: {e}")
    
    def format_summary(self, summary: str, topic: Optional[str] = None, 
                      news_count: Optional[int] = None, sources: Optional[List[str]] = None) -> str:
        """
        Форматирование сводки для отправки пользователю согласно гайду
        
        Args:
            summary: Текст сводки (уже содержит структуру согласно промпту)
            topic: Тема сводки (опционально)
            news_count: Количество обработанных новостей (опционально)
            sources: Список источников (опционально)
        
        Returns:
            Отформатированный текст согласно гайду
        """
        # Начинаем с хэштега согласно гайду
        formatted = "#новости\n\n"
        
        # Добавляем саму сводку (структура: заголовок+новость+источник, инфоповод, банковские решения, инфоповод для звонка/small talk, предложение/мостик)
        formatted += summary
        
        # Добавляем дополнительную информацию в конце, если есть
        if news_count or sources:
            formatted += "\n\n"
            if news_count:
                formatted += f"📰 Обработано новостей: {news_count}\n"
            if sources:
                formatted += f"📡 Источники: {', '.join(sources)}"
        
        return formatted
    
    def get_news_statistics(self) -> str:
        """
        Получение статистики по новостям в текстовом формате
        
        Returns:
            Текстовая статистика
        """
        stats = self.collector.get_statistics()
        
        text = "📈 СТАТИСТИКА ПО НОВОСТЯМ\n\n"
        text += f"Всего новостей в базе: {stats['total']}\n"
        text += f"Новостей за последние 24 часа: {stats['last_24h']}\n\n"
        
        if stats['by_source']:
            text += "Топ источников за последние 24 часа:\n"
            for i, (source, count) in enumerate(stats['by_source'].items(), 1):
                text += f"{i}. {source}: {count} новостей\n"
        
        return text